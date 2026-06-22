import streamlit as st
import pandas as pd
from openpyxl import load_workbook

st.set_page_config(page_title="Indent Summary Generator")

st.title("Indent Summary Generator")

uploaded_file = st.file_uploader(
    "Upload Excel File",
    type=["xlsx"]
)

if uploaded_file is not None:

    try:

        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active

        df = pd.read_excel(uploaded_file)

        latest_date = pd.to_datetime(df["Created On"]).dt.date.max()

        tpt_docs = set()
        dot_docs = set()
        dealer_docs = set()

        tpt_after_1400 = set()
        dot_after_1400 = set()

        mg_qty = 0
        hsd_qty = 0

        for row in range(2, ws.max_row + 1):

            sales_doc = ws.cell(row, 5).value
            material = str(ws.cell(row, 14).value)
            qty = ws.cell(row, 15).value

            if qty is None:
                qty = 0

            if material in ["16730", "17295"]:
                mg_qty += qty

            if material in ["50700", "50800"]:
                hsd_qty += qty

            fill = ws.cell(row, 1).fill

            color = None

            if fill.fill_type == "solid":
                color = fill.fgColor.indexed

            if fill.fill_type is None:
                tpt_docs.add(sales_doc)

            elif color == 47:
                dot_docs.add(sales_doc)

            elif color == 50:
                dealer_docs.add(sales_doc)

            try:

                row_date = pd.to_datetime(
                    ws.cell(row, 21).value
                ).date()

                hour = pd.to_datetime(
                    str(ws.cell(row, 22).value)
                ).hour

                if row_date == latest_date and hour >= 14:

                    if fill.fill_type is None:
                        tpt_after_1400.add(sales_doc)

                    elif color == 47:
                        dot_after_1400.add(sales_doc)

            except:
                pass

        total = (
            len(tpt_docs)
            + len(dot_docs)
            + len(dealer_docs)
        )

        st.text(f"""
Indent Summary Report
---------------------

TPT Indents                : {len(tpt_docs)}
(TPT Indents after 14:00   : {len(tpt_after_1400)})

DoT Indents                : {len(dot_docs)}
(DoT Indents after 14:00   : {len(dot_after_1400)})

Dealer Indents             : {len(dealer_docs)}

Total Indents              : {total}

MG Qty (16730+17295)       : {mg_qty:.0f} KL

HSD Qty (50700+50800)      : {hsd_qty:.0f} KL
""")

    except Exception as e:

        st.error(str(e))
